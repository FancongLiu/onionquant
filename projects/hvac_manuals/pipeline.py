"""
HVAC PDF Manual Search Pipeline
================================
Complete workflow:
  1. Load Excel → group rows into unique manuals (~179 from 482 rows)
  2. Search structured sources per brand (sitemap → product page → DocumentURL → CDN)
  3. Download PDFs with proper naming
  4. Verify PDF content against Excel indicators (pdfplumber)
  5. Write results back to Excel with Match_Status column

Verification is script-based (pdfplumber), not AI. Only edge cases get flagged.
"""

import sys
import io
import re
import json
import subprocess
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timezone

import openpyxl
import requests

# --- Paths ---
EXCEL_PATH = Path(r"E:\HVAC_PDF_search\Unitary_pdf_manual_search_0519_lite.xlsx")
PDF_DIR = Path(__file__).parent / "pdf_downloads"
MANIFEST_PATH = Path(__file__).parent / "manifest.json"
RESULTS_CSV = Path(__file__).parent / "manual_search_results.csv"

PDF_DIR.mkdir(parents=True, exist_ok=True)

# --- Brand domain / sitemap / CDN configuration ---
BRAND_CONFIG = {
    "RHEEM": {
        "domains": ["rheem.com", "ruud.com"],
        "sitemaps": [
            "https://www.rheem.com/wp-sitemap-products-1.xml",
            "https://www.rheem.com/wp-sitemap-products-2.xml",
            "https://www.rheem.com/wp-sitemap-products-3.xml",
        ],
        "cdn_base": "https://files.myrheem.com/webpartners/ProductDocuments/",
        "product_url_pattern": "/products/",
        "document_key": "DocumentURL",
        "related_brands": ["RUUD"],
    },
    "RUUD": {
        "domains": ["ruud.com", "rheem.com"],
        "sitemaps": ["https://www.ruud.com/wp-sitemap-products-1.xml"],
        "cdn_base": "https://files.myrheem.com/webpartners/ProductDocuments/",
        "product_url_pattern": "/products/",
        "document_key": "DocumentURL",
        "related_brands": ["RHEEM"],
    },
    "CARRIER": {
        "domains": ["carrier.com", "carriercca.com"],
        "sitemaps": [],
        "cdn_base": "https://carriercca.com/pdf/products_pdf/",
        "product_url_pattern": "/commercial/en/us/products/",
        "document_key": "pdf",
        "related_brands": ["BRYANT"],
    },
    "BRYANT": {
        "domains": ["bryant.com", "carrier.com"],
        "sitemaps": [],
        "cdn_base": "",
        "product_url_pattern": "/products/",
        "document_key": "DocumentURL",
        "related_brands": ["CARRIER"],
    },
    "LENNOX": {
        "domains": ["lennox.com", "lennoxcommercial.com", "lennoxpros.com"],
        "sitemaps": [],
        "cdn_base": "https://lennox.com/dA/",
        "product_url_pattern": "/products/",
        "document_key": "DocumentURL",
        "related_brands": [],
    },
    "BOSCH": {
        "domains": ["bosch-homecomfort.us", "bosch-thermotechnology.us"],
        "sitemaps": [],
        "cdn_base": "",
        "product_url_pattern": "/products/",
        "document_key": "DocumentURL",
        "related_brands": [],
    },
    "YORK": {
        "domains": ["york.com", "johnsoncontrols.com"],
        "sitemaps": [],
        "cdn_base": "",
        "product_url_pattern": "/products/",
        "document_key": "DocumentURL",
        "related_brands": ["JOHNSON CONTROLS"],
    },
    "JOHNSON CONTROLS": {
        "domains": ["johnsoncontrols.com", "docs.johnsoncontrols.com"],
        "sitemaps": [],
        "cdn_base": "https://docs.johnsoncontrols.com/ductedsystems/",
        "hvac_navigator": "https://files.hvacnavigator.com/p/",
        "khub_api": "https://docs.johnsoncontrols.com/ductedsystems/api/khub/documents",
        "product_url_pattern": "/ductedsystems/",
        "document_key": "DocumentURL",
        "related_brands": ["YORK"],
    },
}

# Request session with retry
session = requests.Session()
session.headers.update(
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }
)


def _ts():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


# ═══════════════════════════════════════════════════════════════
# PHASE 1: Load & Group
# ═══════════════════════════════════════════════════════════════


def load_excel():
    """Load Excel, return (rows, headers, workbook, worksheet)."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["sheet1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows, headers, wb, ws


def model_series(model: str) -> str:
    """Extract alphabetic prefix: 'AD15' -> 'AD', 'RACG2T180AC' -> 'RACG'."""
    if not model:
        return "UNKNOWN"
    m = re.match(r"^([A-Z]+)", str(model))
    return m.group(1) if m else str(model)[:8]


def model_token(model: str) -> str:
    """Extract searchable token: 'RACG2T180AC' -> 'RACG2T', 'AD15' -> 'AD15'."""
    if not model:
        return "UNKNOWN"
    m = re.match(r"^([A-Z]+\d+[A-Z]?)", str(model))
    return m.group(1) if m else model[:10]


def normalize_brand(brand_label: str) -> str:
    """Normalize brand label to canonical name."""
    if not brand_label:
        return "UNKNOWN"
    b = brand_label.upper()
    if "RHEEM" in b and "RUUD" not in b:
        return "RHEEM"
    if "RUUD" in b:
        return "RUUD"
    if "CARRIER" in b:
        return "CARRIER"
    if "BRYANT" in b:
        return "BRYANT"
    if "LENNOX" in b:
        return "LENNOX"
    if "BOSCH" in b:
        return "BOSCH"
    if "YORK" in b:
        return "YORK"
    if "JOHNSON" in b:
        return "JOHNSON CONTROLS"
    return b


def build_manifest(rows):
    """Group rows into unique manuals. Returns (manifest, row_to_manual)."""
    if MANIFEST_PATH.exists():
        with open(MANIFEST_PATH, "r", encoding="utf-8") as f:
            manifest = json.load(f)
    else:
        manifest = {}

    row_to_manual = {}

    for i, r in enumerate(rows):
        brand = normalize_brand(str(r.get("Brand_Scope_Label", "")))
        model = str(r.get("ModelNumberUle", ""))
        query = str(r.get("Suggested_Manual_Query", ""))
        series = model_series(model)
        ref_id = str(r.get("ReferenceId", ""))
        pdf_link_col = str(r.get("pdf_link", "")).strip() if r.get("pdf_link") else ""

        # Deduplicate: brand + series = one manual
        manual_id = f"{brand}__{series}"

        if manual_id not in manifest:
            manifest[manual_id] = {
                "manual_id": manual_id,
                "brand": brand,
                "series": series,
                "pdf_filename": None,
                "pdf_url": None,
                "status": "searching",
                "verdict": None,
                "evidence": "",
                "row_indices": [],
                "queries": [],
                "models": [],
                "reference_ids": [],
                "capacities": [],
                "refrigerants": [],
            }

        entry = manifest[manual_id]
        entry["row_indices"].append(i)
        if query and query not in entry["queries"]:
            entry["queries"].append(query)
        if model and model not in entry["models"]:
            entry["models"].append(model)
        if ref_id and ref_id not in entry["reference_ids"]:
            entry["reference_ids"].append(ref_id)

        cap = str(r.get("CoolingCapacity95FSearchULE", ""))
        if cap and cap not in entry["capacities"]:
            entry["capacities"].append(cap)
        ref = str(r.get("RefrigerantTypeUle", ""))
        if ref and ref not in entry["refrigerants"]:
            entry["refrigerants"].append(ref)

        row_to_manual[i] = manual_id

    return manifest, row_to_manual


# ═══════════════════════════════════════════════════════════════
# PHASE 2: Search & Download
# ═══════════════════════════════════════════════════════════════


def download_pdf(url: str, output_path: Path, timeout: int = 120) -> bool:
    """Download a PDF from URL. Returns True if valid PDF downloaded."""
    try:
        result = subprocess.run(
            [
                "curl",
                "-s",
                "-L",
                "-o",
                str(output_path),
                "-w",
                "%{http_code}",
                "--max-time",
                str(timeout),
                "-H",
                "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                url,
            ],
            capture_output=True,
            text=True,
            timeout=timeout + 10,
        )
        http_code = result.stdout.strip()
        if (
            http_code == "200"
            and output_path.exists()
            and output_path.stat().st_size > 1000
        ):
            with open(output_path, "rb") as f:
                header = f.read(5)
            if header.startswith(b"%PDF"):
                return True
            output_path.unlink()
        elif output_path.exists() and output_path.stat().st_size < 1000:
            output_path.unlink()
        return False
    except Exception as e:
        print(f"  [ERROR] Download failed: {url[:100]} — {e}")
        return False


def fetch_url(url: str, timeout: int = 30) -> str | None:
    """Fetch a URL and return text content."""
    try:
        r = session.get(url, timeout=timeout)
        if r.status_code == 200:
            return r.text
        return None
    except Exception as e:
        print(f"  [ERROR] Fetch failed: {url[:100]} — {e}")
        return None


def extract_pdf_links_from_html(html: str, brand: str) -> list[dict]:
    """Extract DocumentURL / PDF links from product page HTML."""
    results = []
    cfg = BRAND_CONFIG.get(brand, {})

    # Pattern 1: DocumentURL JSON-like embedded data (Rheem style)
    # "DocumentURL":"https://files.myrheem.com/..."
    doc_urls = re.findall(r'"DocumentURL"\s*:\s*"(https?://[^"]+\.pdf)"', html)
    for url in doc_urls:
        results.append({"url": url, "source": "DocumentURL_json"})

    # Pattern 2: documentType + DocumentURL pairs
    doc_type_matches = re.findall(
        r'"documentType"\s*:\s*"([^"]+)"[^}]+"DocumentURL"\s*:\s*"(https?://[^"]+\.pdf)"',
        html,
        re.DOTALL,
    )
    for dtype, url in doc_type_matches:
        results.append({"url": url, "source": f"DocumentURL_{dtype}"})

    # Pattern 3: Generic href to .pdf
    pdf_hrefs = re.findall(r'href\s*=\s*"(https?://[^"]+\.pdf)"', html)
    for url in pdf_hrefs:
        if url not in [r["url"] for r in results]:
            results.append({"url": url, "source": "href_pdf"})

    # Pattern 4: data-url or data-src with .pdf
    data_pdfs = re.findall(r'data-(?:url|src)\s*=\s*"(https?://[^"]+\.pdf)"', html)
    for url in data_pdfs:
        if url not in [r["url"] for r in results]:
            results.append({"url": url, "source": "data_attr"})

    return results


def find_sitemap_urls(brand: str) -> list[str]:
    """Try common sitemap URL patterns for a brand."""
    cfg = BRAND_CONFIG.get(brand, {})
    urls = list(cfg.get("sitemaps", []))

    # Try common WordPress sitemap patterns
    for domain in cfg.get("domains", []):
        for pattern in [
            f"https://www.{domain}/wp-sitemap-products-1.xml",
            f"https://www.{domain}/wp-sitemap.xml",
            f"https://www.{domain}/sitemap.xml",
            f"https://www.{domain}/products-sitemap.xml",
            f"https://www.{domain}/page-sitemap.xml",
            f"https://www.{domain}/wp-sitemap-posts-product-1.xml",
        ]:
            if pattern not in urls:
                urls.append(pattern)
    return urls


def match_sitemap_to_models(
    sitemap_xml: str, models: list[str], brand: str
) -> list[dict]:
    """Match sitemap product URLs to model tokens."""
    matches = []
    tokens = [model_token(m) for m in models if model_token(m) != "UNKNOWN"]
    cfg = BRAND_CONFIG.get(brand, {})

    # Extract URLs from sitemap XML
    all_urls = re.findall(r"<loc>([^<]+)</loc>", sitemap_xml)

    for url in all_urls:
        url_upper = url.upper()
        for token in tokens:
            if token.upper() in url_upper:
                # Prefer product pages over generic pages
                if cfg.get("product_url_pattern", "/products/") in url.lower():
                    matches.append({"url": url, "token": token, "score": 2})
                else:
                    matches.append({"url": url, "token": token, "score": 1})
                break

    # Deduplicate by URL, keeping highest score
    seen = {}
    for m in matches:
        if m["url"] not in seen or m["score"] > seen[m["url"]]["score"]:
            seen[m["url"]] = m
    return sorted(seen.values(), key=lambda x: -x["score"])


def search_cdn_by_model(brand: str, model: str) -> list[str]:
    """Search CDN paths for model-specific PDFs."""
    cfg = BRAND_CONFIG.get(brand, {})
    results = []

    cdn_base = cfg.get("cdn_base", "")
    if not cdn_base:
        return results

    token = model_token(model)
    # Common PDF naming patterns on CDNs
    patterns = [
        f"{token}.",
        f"{token.upper()}.",
        f"{token.lower()}.",
        f"{token}_",
        f"{model_series(model)}-",
        f"{model_series(model)}_",
    ]

    # We can't list CDN directories, but we can try direct guesses
    for pattern in patterns:
        # Try common HVAC manual naming patterns
        for suffix in [
            "product_data.pdf",
            "technical_guide.pdf",
            "spec_sheet.pdf",
            "ProductData.pdf",
            "TechnicalGuide.pdf",
            "SpecSheet.pdf",
            "pd.pdf",
            "tg.pdf",
            "ss.pdf",
        ]:
            results.append(f"{cdn_base}{pattern}{suffix}")

    return results


# ═══════════════════════════════════════════════════════════════
# PHASE 3: PDF Verification (pdfplumber)
# ═══════════════════════════════════════════════════════════════


def extract_pdf_text(pdf_path: Path, max_pages: int = 60) -> str:
    """Extract full text from a PDF, limited to first N pages."""
    try:
        import pdfplumber

        texts = []
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages[:max_pages]:
                t = page.extract_text()
                if t:
                    texts.append(t)
        return "\n".join(texts)
    except Exception as e:
        print(f"  [ERROR] PDF extraction failed: {pdf_path.name} — {e}")
        return ""


def verify_pdf_against_row(pdf_text: str, row: dict) -> dict:
    """
    Verify a PDF against a single Excel row.
    Returns verdict, score, and evidence dict.
    """
    brand = normalize_brand(str(row.get("Brand_Scope_Label", "")))
    model_full = str(row.get("ModelNumberUle", "")).strip()
    model_tok = model_token(model_full)
    series = model_series(model_full)
    ref_type = str(row.get("RefrigerantTypeUle", "")).strip()
    capacity = str(row.get("CoolingCapacity95FSearchULE", "")).strip()
    eer = str(row.get("EER95FSearchULE", "")).strip()
    ieer = str(row.get("IEERSearchUle", "")).strip()
    ahri_type = str(row.get("AHRIType", "")).strip()

    evidence_parts = []
    score = 0
    max_score = 25  # brand(5) + model(5) + refrigerant(5) + capacity(3) + EER(3) + IEER(3) + bonus(1)

    text_upper = pdf_text.upper()

    # -- Brand match (5 pts) --
    brand_matched = False
    brand_aliases = {
        "RHEEM": ["RHEEM", "RUUD"],
        "RUUD": ["RUUD", "RHEEM"],
        "CARRIER": ["CARRIER", "CARRIER®", "BRYANT"],
        "BRYANT": ["BRYANT", "CARRIER"],
        "LENNOX": ["LENNOX", "LENNOX®"],
        "BOSCH": ["BOSCH", "BOSCH THERMOTECHNOLOGY"],
        "YORK": ["YORK", "YORK®", "JOHNSON CONTROLS"],
        "JOHNSON CONTROLS": ["JOHNSON CONTROLS", "JCI", "YORK", "CHOICE"],
    }
    for alias in brand_aliases.get(brand, [brand]):
        if alias in text_upper:
            brand_matched = True
            evidence_parts.append(f"Brand: {alias}")
            score += 5
            break
    if not brand_matched:
        evidence_parts.append(f"Brand NOT found: expected {brand}")

    # -- Model match (5 pts) --
    model_matched = False
    # Try exact model
    if model_full and model_full.upper() in text_upper:
        model_matched = True
        score += 5
        evidence_parts.append(f"Model: {model_full} (exact)")
    elif model_tok and model_tok.upper() in text_upper:
        model_matched = True
        score += 4
        evidence_parts.append(f"Model: {model_tok} (token)")
    elif series and series.upper() in text_upper:
        model_matched = True
        score += 2
        evidence_parts.append(f"Model: {series} (series only)")
    else:
        evidence_parts.append(f"Model NOT found: {model_full}")

    # -- Refrigerant match (5 pts) --
    ref_matched = False
    if ref_type:
        ref_variants = [ref_type, ref_type.replace("-", ""), ref_type.replace("-", " ")]
        for rv in ref_variants:
            if rv.upper() in text_upper:
                ref_matched = True
                score += 5
                evidence_parts.append(f"Refrigerant: {ref_type}")
                break
    if not ref_matched:
        evidence_parts.append(f"Refrigerant: {ref_type} NOT confirmed")

    # -- Capacity match (3 pts) —
    cap_matched = False
    if capacity:
        # Extract first number from capacity (e.g., "172000/172000" -> 172000)
        cap_nums = re.findall(r"(\d+)", capacity)
        for cn in cap_nums[:2]:
            if cn in pdf_text:
                cap_matched = True
                score += 3
                evidence_parts.append(f"Capacity: {cn} BTU")
                break
    if not cap_matched:
        evidence_parts.append(f"Capacity: {capacity} NOT confirmed")

    # -- EER match (3 pts) —
    eer_matched = False
    if eer:
        eer_prefix = eer.split("/")[0]
        # Allow near-match (e.g., "11.00" matches "11.0", "11.0 0")
        eer_short = eer_prefix[:4]
        if eer_prefix in pdf_text or eer_short in pdf_text:
            eer_matched = True
            score += 3
            evidence_parts.append(f"EER: {eer_prefix}")
    if not eer_matched:
        evidence_parts.append(f"EER: {eer} NOT confirmed")

    # -- IEER match (3 pts) —
    ieer_matched = False
    if ieer:
        ieer_prefix = ieer.split("/")[0]
        ieer_short = ieer_prefix[:4]
        if ieer_prefix in pdf_text or ieer_short in pdf_text:
            ieer_matched = True
            score += 3
            evidence_parts.append(f"IEER: {ieer_prefix}")
    if not ieer_matched:
        evidence_parts.append(f"IEER: {ieer} NOT confirmed")

    # -- Bonus: AHRI type match (1 pt) —
    if ahri_type and ahri_type.upper() in text_upper:
        score += 1
        evidence_parts.append(f"Type: {ahri_type}")

    # -- Verdict --
    if score >= 18:
        verdict = "High confidence"
    elif score >= 10:
        verdict = "Needs human confirmation"
    elif score >= 4:
        verdict = "Low confidence"
    else:
        verdict = "Not found"

    return {
        "verdict": verdict,
        "score": score,
        "evidence": "; ".join(evidence_parts),
        "brand_match": brand_matched,
        "model_match": model_matched,
        "refrigerant_match": ref_matched,
        "capacity_match": cap_matched,
        "eer_match": eer_matched,
        "ieer_match": ieer_matched,
    }


def verify_pdf(pdf_path: Path, rows: list[dict]) -> dict:
    """Full verification of a PDF against its rows. Returns aggregated result."""
    print(f"\n  Verifying: {pdf_path.name} ({pdf_path.stat().st_size / 1024:.0f} KB)")
    pdf_text = extract_pdf_text(pdf_path)

    if not pdf_text or len(pdf_text) < 100:
        return {
            "verdict": "Not found",
            "score": 0,
            "evidence": "PDF text extraction failed or empty",
            "row_results": [],
        }

    row_results = []
    for row in rows:
        result = verify_pdf_against_row(pdf_text, row)
        row_results.append(result)

    # Aggregate: best score across covered rows
    best = max(row_results, key=lambda r: r["score"]) if row_results else {"score": 0}
    avg_score = (
        sum(r["score"] for r in row_results) / len(row_results) if row_results else 0
    )

    return {
        "verdict": best["verdict"],
        "score": best["score"],
        "avg_score": round(avg_score, 1),
        "evidence": best["evidence"],
        "row_results": row_results,
    }


# ═══════════════════════════════════════════════════════════════
# PHASE 4: Results Output
# ═══════════════════════════════════════════════════════════════


def write_results_to_excel(manifest, row_to_manual):
    """Write pdf_filename, pdf_link, Comments, and Match_Status back to Excel."""
    _, headers, wb, ws = load_excel()

    # Ensure output columns exist
    extra_cols = ["pdf_filename", "Match_Status", "Score", "Evidence"]
    for col_name in extra_cols:
        if col_name not in headers:
            col_idx = len(headers) + 1
            ws.cell(row=1, column=col_idx, value=col_name)
            headers.append(col_name)

    pdf_col = headers.index("pdf_filename") + 1
    match_col = headers.index("Match_Status") + 1
    score_col = headers.index("Score") + 1
    evidence_col = headers.index("Evidence") + 1
    pdf_link_col = headers.index("pdf_link") + 1

    for row_idx, manual_id in row_to_manual.items():
        entry = manifest.get(manual_id)
        if not entry:
            continue
        excel_row = row_idx + 2

        if entry.get("pdf_filename"):
            ws.cell(row=excel_row, column=pdf_col, value=entry["pdf_filename"])
        if entry.get("verdict"):
            ws.cell(row=excel_row, column=match_col, value=entry["verdict"])
        if entry.get("score") is not None:
            ws.cell(row=excel_row, column=score_col, value=entry["score"])
        if entry.get("evidence"):
            ws.cell(row=excel_row, column=evidence_col, value=entry["evidence"])
        if entry.get("pdf_url"):
            ws.cell(row=excel_row, column=pdf_link_col, value=entry["pdf_url"])

    output_path = EXCEL_PATH.parent / "Unitary_pdf_manual_search_0519_lite_RESULTS.xlsx"
    wb.save(output_path)
    print(f"\nResults written to: {output_path}")
    return output_path


def write_csv_results(manifest):
    """Write detailed results to CSV."""
    import csv

    fieldnames = [
        "ReferenceId",
        "Brand",
        "Series",
        "ModelNumberUle",
        "ExpectedCapacity",
        "ExpectedEER",
        "ExpectedIEER",
        "ExpectedRefrigerant",
        "Verdict",
        "Score",
        "Evidence",
        "PDF_URL",
        "LocalPath",
    ]

    with open(RESULTS_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for mid, entry in sorted(manifest.items()):
            for i, rid in enumerate(entry.get("reference_ids", [entry["series"]])):
                writer.writerow(
                    {
                        "ReferenceId": rid,
                        "Brand": entry["brand"],
                        "Series": entry["series"],
                        "ModelNumberUle": entry["models"][i]
                        if i < len(entry["models"])
                        else "",
                        "ExpectedCapacity": entry["capacities"][i]
                        if i < len(entry["capacities"])
                        else "",
                        "ExpectedEER": "",
                        "ExpectedIEER": "",
                        "ExpectedRefrigerant": entry["refrigerants"][i]
                        if i < len(entry["refrigerants"])
                        else "",
                        "Verdict": entry.get("verdict", ""),
                        "Score": entry.get("score", ""),
                        "Evidence": entry.get("evidence", ""),
                        "PDF_URL": entry.get("pdf_url", ""),
                        "LocalPath": entry.get("pdf_filename", ""),
                    }
                )

    print(f"CSV results written to: {RESULTS_CSV}")


def print_summary(manifest):
    """Print execution summary."""
    total = len(manifest)
    verdicts = defaultdict(int)
    total_rows = 0
    covered_rows = 0
    for m in manifest.values():
        v = m.get("verdict", "Not found")
        verdicts[v] += 1
        n_rows = len(m.get("reference_ids", [m["series"]]))
        total_rows += n_rows
        if m.get("pdf_filename"):
            covered_rows += n_rows

    print("\n" + "=" * 60)
    print("PIPELINE SUMMARY")
    print("=" * 60)
    print(f"  Unique manuals: {total}")
    print(f"  Total rows:     {total_rows}")
    print(
        f"  PDFs found:     {sum(1 for m in manifest.values() if m.get('pdf_filename'))}"
    )
    print(
        f"  Rows covered:   {covered_rows}/{total_rows} ({covered_rows / max(1, total_rows) * 100:.0f}%)"
    )
    print()
    for v in [
        "High confidence",
        "Needs human confirmation",
        "Low confidence",
        "Not found",
    ]:
        if verdicts[v]:
            print(f"  {v}: {verdicts[v]}")

    # Brand breakdown
    brand_stats = defaultdict(lambda: {"total": 0, "found": 0})
    for m in manifest.values():
        b = m["brand"]
        brand_stats[b]["total"] += 1
        if m.get("pdf_filename"):
            brand_stats[b]["found"] += 1

    print("\n  By Brand:")
    for b, s in sorted(brand_stats.items(), key=lambda x: -x[1]["total"]):
        pct = s["found"] / max(1, s["total"]) * 100
        print(f"    {b}: {s['found']}/{s['total']} PDFs found ({pct:.0f}%)")


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    print("=" * 60)
    print("HVAC PDF Manual Search Pipeline")
    print("=" * 60)

    # Phase 1: Load & Group
    print("\n[Phase 1] Loading Excel and building manifest...")
    rows, headers, wb, ws = load_excel()
    print(f"  Loaded {len(rows)} rows, {len(headers)} columns")

    manifest, row_to_manual = build_manifest(rows)
    print(f"  Grouped into {len(manifest)} unique manuals")

    # Print brand distribution
    brand_counts = defaultdict(int)
    for m in manifest.values():
        brand_counts[m["brand"]] += 1
    for b, c in sorted(brand_counts.items(), key=lambda x: -x[1]):
        print(f"    {b}: {c} manuals")

    # Save manifest
    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    print(f"  Manifest saved: {MANIFEST_PATH}")

    print_summary(manifest)
