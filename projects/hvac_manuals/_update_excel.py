"""Update Excel with score labels, color coding, legend sheet, and per-person stats."""

import json
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

with open(
    "e:/2026_AgentStudy/Python_code/hvac_manuals/manifest.json", encoding="utf-8"
) as f:
    manifest = json.load(f)

ref_to_entry = {}
for key, entry in manifest.items():
    for ref_id in entry.get("reference_ids", []):
        ref_to_entry[str(ref_id)] = entry

wb = openpyxl.load_workbook("E:/HVAC_PDF_search/Unitary_pdf_manual_search_RESULTS.xlsx")
ws = wb.active

# Read headers
headers = {}
for c in range(1, ws.max_column + 1):
    headers[ws.cell(1, c).value] = c

ref_id_col = headers["ReferenceId"]
comments_col = 3
status_col = 18
verdict_col = 19
score_col = 20
series_id_col = 21

# Add score_label column
score_label_col = 22
ws.cell(1, score_label_col, "score_label")
ws.cell(1, score_label_col).font = Font(bold=True)

# Fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
orange_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def score_info(score):
    if score >= 13:
        return (
            "EXACT",
            "Same brand + series + refrigerant. Capacity range matches. Direct document.",
            green_fill,
        )
    if score >= 10:
        return (
            "HIGH",
            "Same product family + correct refrigerant. Doc covers this model family line.",
            green_fill,
        )
    if score >= 6:
        return (
            "MEDIUM",
            "Same product architecture, different sub-series or inferred match.",
            yellow_fill,
        )
    if score >= 4:
        return (
            "LOW",
            "Same product family but WRONG refrigerant. R-410A doc for R-454B product. Performance tables will differ.",
            orange_fill,
        )
    if score >= 3:
        return (
            "LEGACY",
            "Approximate match, low confidence. NEEDS HUMAN REVIEW.",
            red_fill,
        )
    return (
        "NOT_FOUND",
        "No public documentation found through any channel (API/CDN/sitemap/search).",
        red_fill,
    )


# Process rows
team_stats = defaultdict(
    lambda: {
        "total": 0,
        "found": 0,
        "exact": 0,
        "high": 0,
        "medium": 0,
        "low": 0,
        "legacy": 0,
        "not_found": 0,
    }
)

for row in range(2, ws.max_row + 1):
    ref_id = str(ws.cell(row, ref_id_col).value or "")
    entry = ref_to_entry.get(ref_id, {})
    status = entry.get("status", "searching")
    score = entry.get("score", 0)
    team = str(ws.cell(row, headers["Team_Label"]).value or "")

    label, desc, fill = score_info(score)

    # Write score_label
    ws.cell(row, score_label_col, label)
    ws.cell(row, score_label_col).fill = fill

    # Rewrite Comments with clearer format
    evidence = entry.get("evidence", "")
    pdf_name = entry.get("pdf_filename", "N/A")
    comment = f"[{label}] {desc} | PDF: {pdf_name}"
    if len(comment) < 500:
        comment += f" | Evidence: {evidence[:200]}"
    ws.cell(row, comments_col, comment[:500])

    # Color status and score columns
    ws.cell(row, status_col).fill = fill
    ws.cell(row, score_col).fill = fill

    # Stats
    team_stats[team]["total"] += 1
    if status == "found":
        team_stats[team]["found"] += 1
        if score >= 13:
            team_stats[team]["exact"] += 1
        elif score >= 10:
            team_stats[team]["high"] += 1
        elif score >= 6:
            team_stats[team]["medium"] += 1
        elif score >= 4:
            team_stats[team]["low"] += 1
        else:
            team_stats[team]["legacy"] += 1
    else:
        team_stats[team]["not_found"] += 1

# Adjust column widths
ws.column_dimensions[get_column_letter(comments_col)].width = 90
ws.column_dimensions[get_column_letter(score_label_col)].width = 14

# ============ Legend Sheet ============
ws2 = wb.create_sheet("Score Legend & Notes")

legend = [
    ("SCORE LEGEND — HVAC Manual Search Results", True, 14, None),
    ("", False, None, None),
    ("How to read the 'score_label' and 'score' columns:", False, None, None),
    ("", False, None, None),
    ("Score", True, None, None),
    ("13 = EXACT", False, None, green_fill),
    (
        "    Same brand + same series + correct refrigerant. Capacity range matches. This is the definitive document for this product.",
        False,
        None,
        None,
    ),
    ("    Action: None. Ready to use.", False, None, None),
    ("", False, None, None),
    ("10 = HIGH", False, None, green_fill),
    (
        "    Same product family + correct refrigerant. Doc covers this product line but may be titled for a different sub-series.",
        False,
        None,
        None,
    ),
    (
        "    Example: JH series uses Select JV28-JV50 Tech Guide. Both are Select product line, same refrigerant.",
        False,
        None,
        None,
    ),
    (
        "    Action: None. Document is correct. Just verify the specific model capacity is in the performance tables.",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("6 = MEDIUM", False, None, yellow_fill),
    (
        "    Same product architecture but different sub-series or inferred match. Doc for sibling series, same OEM.",
        False,
        None,
        None,
    ),
    (
        "    Example: ZLE series uses Core ZX/ZY/ZQ/ZL Tech Guide. Both are Core line but different sub-series. Also has Bosch OEM counterpart.",
        False,
        None,
        None,
    ),
    (
        "    Action: Review doc to confirm it covers this specific model. Likely correct but worth double-checking.",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("4 = LOW", False, None, orange_fill),
    (
        "    Same product family but WRONG REFRIGERANT generation. R-410A doc used for R-454B product (or vice versa).",
        False,
        None,
        None,
    ),
    (
        "    Product architecture is identical but refrigerant parameters (pressure, capacity, EER) will differ.",
        False,
        None,
        None,
    ),
    (
        "    Example: VH series (R-454B) uses Select JV28-JV50 Tech Guide (R-410A). Architecture same, performance tables different.",
        False,
        None,
        None,
    ),
    (
        "    Action: Use with caution. Performance data needs refrigerant-specific adjustment. Best available until new docs published.",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("3 = LEGACY", False, None, red_fill),
    (
        "    Approximate/legacy match. Low confidence guess based on limited model info. May be completely wrong.",
        False,
        None,
        None,
    ),
    (
        "    Action: NEEDS HUMAN REVIEW. Try to find a better match or confirm from manufacturer directly.",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("0 = NOT_FOUND", False, None, red_fill),
    (
        "    No public documentation found through any channel: API, CDN, sitemap, OEM cross-reference, web search.",
        False,
        None,
        None,
    ),
    (
        "    Common reasons: legacy/discontinued model, commercial split system (never publicly documented), very new refrigerant generation.",
        False,
        None,
        None,
    ),
    (
        "    Action: Try archive.org, manualslib.com, or dealer portal if available.",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("---", False, None, None),
    ("", False, None, None),
    ("NOTES FOR THE NEXT AI / RESEARCHER", True, 12, None),
    ("", False, None, None),
    (
        "1. All 97 NOT_FOUND rows have been searched through ALL known public channels:",
        False,
        None,
        None,
    ),
    ("   - Manufacturer sitemaps (Rheem, Carrier, Lennox)", False, None, None),
    ("   - Public CDN (files.myrheem.com, files.hvacnavigator.com)", False, None, None),
    (
        "   - JCI Fluid Topics API (docs.johnsoncontrols.com/ductedsystems/api/khub/)",
        False,
        None,
        None,
    ),
    ("   - Web search with exact model numbers", False, None, None),
    (
        "   - OEM cross-reference (Bosch=JCI, Bryant=Carrier, Ruud=Rheem)",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("2. The 42 not_found series fall into 3 categories:", False, None, None),
    (
        "   a) Legacy/discontinued models removed from manufacturer websites, no public archive",
        False,
        None,
        None,
    ),
    (
        "   b) Commercial split systems (KC/KD/KE/WC/WD/WE) never publicly documented",
        False,
        None,
        None,
    ),
    (
        "   c) Very new R-454B models (VH/VV/VX/VY) docs not yet published",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("3. Breakthrough approaches for remaining items:", False, None, None),
    (
        "   - archive.org / web.archive.org for historical product page snapshots",
        False,
        None,
        None,
    ),
    (
        "   - manualslib.com, manualsbrain.com, manualzz.com (user-uploaded PDFs)",
        False,
        None,
        None,
    ),
    (
        "   - HVAC distributor websites (may still host PDFs removed from manufacturer sites)",
        False,
        None,
        None,
    ),
    (
        "   - Contact manufacturer technical support directly for legacy model documentation",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    ("4. JCI Fluid Topics API (public, NO login required):", False, None, None),
    (
        "   Search: POST https://docs.johnsoncontrols.com/ductedsystems/api/khub/clustered-search",
        False,
        None,
        None,
    ),
    ('   Body: {"search": "<model number>"}', False, None, None),
    (
        "   Download PDF: GET https://docs.johnsoncontrols.com/ductedsystems/api/khub/documents/{documentId}/content",
        False,
        None,
        None,
    ),
    (
        "   Full doc list (1208): GET https://docs.johnsoncontrols.com/ductedsystems/api/khub/documents?search=&size=1208",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    (
        "5. See also: HANDOFF_REMAINING_NOT_FOUND.md for full per-series breakdown.",
        False,
        None,
        None,
    ),
    (
        "   Location: E:/HVAC_PDF_search/HANDOFF_REMAINING_NOT_FOUND.md",
        False,
        None,
        None,
    ),
    ("", False, None, None),
    (
        "Generated: 2026-05-26 | Total: 482 rows, 385 found (80%), 97 not_found (20%)",
        False,
        None,
        None,
    ),
]

for r, (text, is_bold, font_size, fill) in enumerate(legend, 1):
    cell = ws2.cell(r, 1, text)
    if is_bold and font_size:
        cell.font = Font(bold=True, size=font_size)
    elif is_bold:
        cell.font = Font(bold=True)
    if fill:
        cell.fill = fill

ws2.column_dimensions["A"].width = 120

# ============ Per-Person Stats Sheet ============
ws3 = wb.create_sheet("Per-Person Stats")

stats_h = [
    "Person",
    "Total",
    "Found",
    "%",
    "EXACT(13)",
    "HIGH(10)",
    "MEDIUM(6)",
    "LOW(4)",
    "LEGACY(3)",
    "NOT_FOUND(0)",
]
for c, h in enumerate(stats_h, 1):
    cell = ws3.cell(1, c, h)
    cell.font = Font(bold=True)

# Legend for stat columns
ws3.cell(1, 5).fill = green_fill
ws3.cell(1, 6).fill = green_fill
ws3.cell(1, 7).fill = yellow_fill
ws3.cell(1, 8).fill = orange_fill
ws3.cell(1, 9).fill = red_fill
ws3.cell(1, 10).fill = red_fill

row_idx = 2
for team in sorted(team_stats.keys()):
    d = team_stats[team]
    pct = d["found"] / d["total"] * 100 if d["total"] > 0 else 0
    vals = [
        team,
        d["total"],
        d["found"],
        f"{pct:.0f}%",
        d["exact"],
        d["high"],
        d["medium"],
        d["low"],
        d["legacy"],
        d["not_found"],
    ]
    for c, val in enumerate(vals, 1):
        cell = ws3.cell(row_idx, c, val)
        if c >= 5:
            # Color by score tier
            if c == 5 or c == 6:
                cell.fill = green_fill
            elif c == 7:
                cell.fill = yellow_fill
            elif c == 8:
                cell.fill = orange_fill
            elif c == 9 or c == 10:
                cell.fill = red_fill
    row_idx += 1

# Totals
totals = ["TOTAL"]
for c in range(2, len(stats_h) + 1):
    total = sum(
        ws3.cell(r, c).value
        for r in range(2, row_idx)
        if isinstance(ws3.cell(r, c).value, (int, float))
    )
    totals.append(total)
for c, val in enumerate(totals, 1):
    cell = ws3.cell(row_idx, c, val)
    cell.font = Font(bold=True)
    if c == 3 or c == 4:
        cell.font = Font(bold=True)

ws3.column_dimensions["A"].width = 22
for c in range(2, len(stats_h) + 1):
    ws3.column_dimensions[get_column_letter(c)].width = 16

# Save
output_path = "E:/HVAC_PDF_search/Unitary_pdf_manual_search_RESULTS.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print("3 sheets: Data + Score Legend & Notes + Per-Person Stats")

# Print summary
print("\nPer-person breakdown:")
for team in sorted(team_stats.keys()):
    d = team_stats[team]
    pct = d["found"] / d["total"] * 100 if d["total"] > 0 else 0
    print(
        f"  {team}: {d['found']}/{d['total']} ({pct:.0f}%) "
        f"E={d['exact']} H={d['high']} M={d['medium']} L={d['low']} LG={d['legacy']} NF={d['not_found']}"
    )
