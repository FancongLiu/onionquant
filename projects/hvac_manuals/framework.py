"""
HVAC Manual Search & Download Framework
========================================
- Groups 482 Excel rows into ~179 unique manuals
- Downloads PDFs from known sources
- Updates Excel with filename mappings
- Tracks progress in manifest.json
"""

import json
import re
import subprocess
from pathlib import Path
from collections import defaultdict

import openpyxl

EXCEL_PATH = Path(r"E:\HVAC_PDF_search\Unitary_pdf_manual_search_0519_lite.xlsx")
PDF_DIR = Path(__file__).parent / "pdf_downloads"
MANIFEST_PATH = Path(__file__).parent / "manifest.json"
MAPPING_PATH = Path(__file__).parent / "manual_map.json"

PDF_DIR.mkdir(parents=True, exist_ok=True)


def load_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["sheet1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows, headers, wb, ws


def build_manifest(rows):
    """
    Build a manifest of unique manuals and which rows they cover.
    A manual is identified by:
    1. pdf_link column (if filled with an actual manual title)
    2. Brand + Model series prefix (e.g., JCI + AD, Bosch + LD)

    Returns: dict of manual_id -> {
        'manual_title': str,
        'brand': str,
        'series': str,
        'pdf_filename': str or None,
        'source_url': str or None,
        'status': 'found' | 'searching' | 'needs_auth' | 'not_found',
        'rows': [row_indices],
        'queries': [Suggested_Manual_Query strings],
        'models': [ModelNumberUle strings],
        'reference_ids': [ReferenceId ints],
    }
    """
    # Load existing manifest if any
    if MANIFEST_PATH.exists():
        with open(MANIFEST_PATH, "r", encoding="utf-8") as f:
            manifest = json.load(f)
    else:
        manifest = {}

    def series_from_model(model):
        """Extract alphabetic prefix from model number: AD15 -> AD, LCT302 -> LCT"""
        if not model:
            return "UNKNOWN"
        m = re.match(r"^([A-Z]+)", str(model))
        return m.group(1) if m else str(model)[:6]

    # Build row-to-manual mapping
    row_manual_map = {}  # row_index -> manual_id

    for i, r in enumerate(rows):
        pdf_link = str(r["pdf_link"]).strip() if r["pdf_link"] else ""
        brand = (
            str(r["Brand_Scope_Label"]).split(" (")[0].strip()
            if r["Brand_Scope_Label"]
            else "UNKNOWN"
        )
        query = str(r["Suggested_Manual_Query"]) if r["Suggested_Manual_Query"] else ""
        model = str(r["ModelNumberUle"]) if r["ModelNumberUle"] else ""

        # Determine manual_id
        if pdf_link and pdf_link not in ("untitled", "/", "None"):
            # Use existing pdf_link title as manual_id (normalized)
            manual_title = pdf_link.split(" • ")[0].strip()
            series = series_from_model(model)
            manual_id = f"{brand}__{series}__{manual_title[:80]}"
        else:
            series = series_from_model(model)
            manual_id = f"{brand}__{series}"

        # Sanitize manual_id
        manual_id = re.sub(r'[<>:"/\\|?*]', "_", manual_id)

        if manual_id not in manifest:
            manifest[manual_id] = {
                "manual_title": pdf_link
                if pdf_link and pdf_link not in ("untitled", "/", "None")
                else None,
                "brand": brand,
                "series": series,
                "pdf_filename": None,
                "source_url": None,
                "status": "searching",
                "rows": [],
                "queries": [],
                "models": [],
                "reference_ids": [],
            }

        entry = manifest[manual_id]
        entry["rows"].append(i)
        if query and query not in entry["queries"]:
            entry["queries"].append(query)
        if model and model not in entry["models"]:
            entry["models"].append(model)
        rid = r["ReferenceId"]
        if rid and rid not in entry["reference_ids"]:
            entry["reference_ids"].append(rid)

        # Pre-fill pf_filename for entries that already have a filename in pdf_link
        if pdf_link and pdf_link.lower().endswith(".pdf"):
            entry["pdf_filename"] = pdf_link
            entry["status"] = "found"

        row_manual_map[i] = manual_id

    return manifest, row_manual_map


def update_excel(manifest, row_manual_map):
    """Write pdf_filename back to Excel rows."""
    _, _, wb, ws = load_excel()

    # Add pdf_filename column if not exists (col 17 = Q)
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "pdf_filename" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="pdf_filename")
        headers.append("pdf_filename")

    pdf_col = headers.index("pdf_filename") + 1  # 1-based

    for row_idx, manual_id in row_manual_map.items():
        entry = manifest.get(manual_id)
        if entry and entry["pdf_filename"]:
            ws.cell(row=row_idx + 2, column=pdf_col, value=entry["pdf_filename"])

    # Also fill pdf_link if status is found
    pdf_link_col = headers.index("pdf_link") + 1
    for row_idx, manual_id in row_manual_map.items():
        entry = manifest.get(manual_id)
        if entry and entry["status"] == "found" and entry.get("manual_title"):
            current = ws.cell(row=row_idx + 2, column=pdf_link_col).value
            if not current or str(current).strip() in ("", "None", "untitled", "/"):
                ws.cell(
                    row=row_idx + 2, column=pdf_link_col, value=entry["manual_title"]
                )

    wb.save(EXCEL_PATH)
    print(f"Excel updated: {EXCEL_PATH}")


def save_manifest(manifest):
    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    print(f"Manifest saved: {MANIFEST_PATH} ({len(manifest)} manuals)")


def attempt_download_jci_direct(document_id: str, output_path: Path) -> bool:
    """Try to download a PDF from Johnson Controls Knowledge Exchange by document ID."""
    url = f"https://docs.johnsoncontrols.com/ductedsystems/api/khub/documents/{document_id}/content"
    try:
        result = subprocess.run(
            ["curl", "-s", "-L", "-o", str(output_path), "-w", "%{http_code}", url],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=120,
        )
        http_code = result.stdout.strip()
        if (
            http_code == "200"
            and output_path.exists()
            and output_path.stat().st_size > 1000
        ):
            return True
        # Clean up failed download
        if output_path.exists():
            output_path.unlink()
        return False
    except Exception as e:
        print(f"  Download error: {e}")
        return False


def attempt_search_manualsnet(query: str) -> list[dict]:
    """Search manualsnet.com for matching PDFs. Returns list of {title, url} dicts."""
    # This requires web scraping - placeholder for now
    return []


def attempt_download_from_url(url: str, output_path: Path) -> bool:
    """Download PDF from a direct URL."""
    try:
        result = subprocess.run(
            [
                "curl",
                "-s",
                "-L",
                "-o",
                str(output_path),
                "-w",
                "%{http_code}",
                "--max-time",
                "120",
                "-H",
                "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                url,
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=130,
        )
        http_code = result.stdout.strip()
        if (
            http_code == "200"
            and output_path.exists()
            and output_path.stat().st_size > 1000
        ):
            # Verify it's actually a PDF
            with open(output_path, "rb") as f:
                header = f.read(5)
            if header.startswith(b"%PDF"):
                return True
            else:
                output_path.unlink()
                return False
        if output_path.exists() and output_path.stat().st_size < 1000:
            output_path.unlink()
        return False
    except Exception as e:
        print(f"  Download error: {e}")
        return False


def print_report(manifest):
    """Print a summary report."""
    total = len(manifest)
    found = sum(1 for m in manifest.values() if m["status"] == "found")
    needs_auth = sum(1 for m in manifest.values() if m["status"] == "needs_auth")
    searching = sum(1 for m in manifest.values() if m["status"] == "searching")
    not_found = sum(1 for m in manifest.values() if m["status"] == "not_found")

    total_rows = sum(len(m["rows"]) for m in manifest.values())
    rows_found = sum(
        len(m["rows"]) for m in manifest.values() if m["status"] == "found"
    )

    print("=" * 70)
    print("MANIFEST REPORT")
    print("=" * 70)
    print(f"  Unique manuals: {total}")
    print(f"  Found:          {found}  ({rows_found}/{total_rows} rows covered)")
    print(f"  Needs auth:     {needs_auth}")
    print(f"  Searching:      {searching}")
    print(f"  Not found:      {not_found}")
    print()

    # Brand breakdown
    brand_stats = defaultdict(lambda: {"total": 0, "found": 0})
    for m in manifest.values():
        brand = m["brand"]
        brand_stats[brand]["total"] += 1
        if m["status"] == "found":
            brand_stats[brand]["found"] += 1

    print("  By Brand:")
    for brand, stats in sorted(brand_stats.items(), key=lambda x: -x[1]["total"]):
        pct = stats["found"] / stats["total"] * 100 if stats["total"] else 0
        print(f"    {brand}: {stats['found']}/{stats['total']} found ({pct:.0f}%)")


if __name__ == "__main__":
    print("Loading Excel...")
    rows, headers, _, _ = load_excel()

    print("Building manifest...")
    manifest, row_manual_map = build_manifest(rows)

    # Mark JCI entries as needs_auth (documents behind login portal)
    for mid, entry in manifest.items():
        if (
            entry["brand"].startswith("JOHNSON CONTROLS")
            and entry["status"] == "searching"
        ):
            entry["status"] = "needs_auth"
            entry["notes"] = (
                "JCI Unitary products require DS Solutions App or dealer login at docs.johnsoncontrols.com/ductedsystems"
            )

    # Mark the example PDF as found
    example_mid = [
        mid
        for mid, m in manifest.items()
        if any("AD15" in q for q in m["queries"]) and "JOHNSON CONTROLS" in m["brand"]
    ]
    for mid in example_mid:
        manifest[mid]["pdf_filename"] = "6411197-jtg-a-1023.pdf"
        manifest[mid]["status"] = "found"
        manifest[mid]["manual_title"] = (
            "Technical Guide: Johnson Controls Choice AD15 to AD28"
        )
        manifest[mid]["source_url"] = (
            "docs.johnsoncontrols.com/ductedsystems (manual download)"
        )

    save_manifest(manifest)
    update_excel(manifest, row_manual_map)
    print_report(manifest)
