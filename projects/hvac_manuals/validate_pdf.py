"""
PDF Validator — extracts specs from downloaded PDFs and matches against Excel rows.
"""

import io
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pdfplumber

PDF_DIR = Path(__file__).parent / "pdf_downloads"
EXCEL_PATH = Path(r"E:\HVAC_PDF_search\Unitary_pdf_manual_search_0519_lite.xlsx")
MANIFEST_PATH = Path(__file__).parent / "manifest.json"


def extract_specs_from_pdf(pdf_path: Path) -> dict[str, list[dict]]:
    """
    Extract performance specs from a PDF.
    Returns: {model_number: [{capacity, eer, ieer, refrigerant, tonnage, ...}]}
    """
    specs = defaultdict(list)

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            # Look for physical data or capacity tables
            # Pattern: model names followed by performance numbers
            # Common patterns in JCI technical guides:
            # "ARI net capacity (BTU)" near model names
            # "EER" rows
            # "IEER" rows

            lines = text.split("\n")
            for i, line in enumerate(lines):
                # Look for lines with model number patterns (e.g., AD15, KD25, etc.)
                model_match = re.findall(r"\b([A-Z]{2,4}\d{2,4}[A-Z]?)\b", line)
                if not model_match:
                    continue

                # Check if nearby lines have capacity/EER/IEER data
                context = "\n".join(lines[max(0, i - 5) : min(len(lines), i + 20)])
                if any(
                    kw in context.lower()
                    for kw in [
                        "capacity",
                        "eer",
                        "ieer",
                        "gross",
                        "net capacity",
                        "btu",
                    ]
                ):
                    for model in model_match:
                        specs[model].append(
                            {
                                "page": page.page_number,
                                "context": context[:500],
                            }
                        )

    return dict(specs)


def extract_physical_data_tables(pdf_path: Path) -> list[dict]:
    """
    Extract structured physical data tables from PDF.
    These tables contain model-by-model specs.
    """
    tables_data = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                # Check if this is a performance/physical data table
                header_text = " ".join([str(c) if c else "" for c in (table[0] or [])])
                if any(
                    kw in header_text.lower()
                    for kw in [
                        "capacity",
                        "eer",
                        "ieer",
                        "tonnage",
                        "nominal",
                        "model",
                        "btu",
                        "gross",
                    ]
                ):
                    tables_data.append(
                        {
                            "page": page.page_number,
                            "table": table,
                            "header": header_text[:200],
                        }
                    )

    return tables_data


def match_specs(pdf_specs: dict, excel_rows: list[dict]) -> list[dict]:
    """
    Match PDF specs against Excel rows. Returns match results with confidence.
    """
    results = []
    for row in excel_rows:
        model_full = str(row["ModelNumberUle"]) if row["ModelNumberUle"] else ""
        # Extract base model from wildcard pattern (e.g., "AD15[C,E]**..." -> "AD15")
        base_model = re.match(r"^([A-Z]+\d+)", model_full)
        model_key = base_model.group(1) if base_model else ""

        excel_capacity = (
            str(row["CoolingCapacity95FSearchULE"]).split("/")[0]
            if row["CoolingCapacity95FSearchULE"]
            else ""
        )
        excel_eer = (
            str(row["EER95FSearchULE"]).split("/")[0] if row["EER95FSearchULE"] else ""
        )
        excel_ieer = (
            str(row["IEERSearchUle"]).split("/")[0] if row["IEERSearchUle"] else ""
        )
        excel_refrigerant = (
            str(row["RefrigerantTypeUle"]).strip() if row["RefrigerantTypeUle"] else ""
        )

        pdf_matches = pdf_specs.get(model_key, [])
        match_found = len(pdf_matches) > 0

        # Check for numeric spec matches in context
        spec_match_score = 0
        if match_found:
            for match in pdf_matches:
                context = match["context"]
                if excel_capacity and excel_capacity in context:
                    spec_match_score += 1
                if excel_eer and excel_eer[:4] in context:
                    spec_match_score += 1
                if excel_refrigerant and excel_refrigerant in context:
                    spec_match_score += 1

        results.append(
            {
                "model": model_key,
                "model_full": model_full[:40],
                "excel_capacity": excel_capacity,
                "excel_eer": excel_eer,
                "excel_ieer": excel_ieer,
                "excel_refrigerant": excel_refrigerant,
                "pdf_found": match_found,
                "spec_match_score": spec_match_score,
                "confidence": "HIGH"
                if spec_match_score >= 2
                else ("MEDIUM" if spec_match_score >= 1 else "LOW"),
            }
        )

    return results


def validate_against_excel(pdf_path: Path, excel_rows: list[dict]) -> dict:
    """Full validation pipeline."""
    print(f"\nValidating: {pdf_path.name}")
    print("-" * 60)

    # Extract physical data tables
    tables = extract_physical_data_tables(pdf_path)
    print(f"  Found {len(tables)} performance data tables")

    # Extract model mentions
    specs = extract_specs_from_pdf(pdf_path)
    models_in_pdf = list(specs.keys())
    print(f"  Models mentioned in PDF: {models_in_pdf[:20]}")

    # Match against Excel
    matches = match_specs(specs, excel_rows)

    matched = [m for m in matches if m["pdf_found"]]
    high_conf = [m for m in matched if m["confidence"] == "HIGH"]

    print(f"  Excel rows matched: {len(matched)}/{len(matches)}")
    print(f"  High confidence: {len(high_conf)}")
    print()

    for m in matched[:10]:
        flag = "✓" if m["confidence"] == "HIGH" else "?"
        print(
            f"  {flag} {m['model']}: cap={m['excel_capacity']}, eer={m['excel_eer']}, "
            f"ieer={m['excel_ieer']}, ref={m['excel_refrigerant']}"
        )

    return {
        "pdf": str(pdf_path.name),
        "tables_found": len(tables),
        "models_in_pdf": models_in_pdf,
        "rows_matched": len(matched),
        "high_confidence": len(high_conf),
        "details": matches,
    }


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["sheet1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        dict(zip(headers, row))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)
    ]

    # Test with the example PDF
    pdf_path = PDF_DIR / "6411197-jtg-a-1023.pdf"
    if pdf_path.exists():
        result = validate_against_excel(pdf_path, rows)
        print(
            f"\n  Summary: {result['high_confidence']}/{result['rows_matched']} high-confidence matches"
        )
