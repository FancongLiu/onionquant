#!/usr/bin/env python3
"""
HVAC PDF Manual Search — Autonomous Overnight Pipeline
=======================================================
Fully unattended: searches structured sources + web fallback,
downloads PDFs, verifies against Excel indicators, writes results.
Checkpoint-resumable — save state after each manual processed.

Usage:
  python auto_pipeline.py          # fresh start
  python auto_pipeline.py --resume # resume from checkpoint
"""

import argparse
import io
import json
import logging
import re
import sys
import time
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import requests
from ddgs import DDGS

# --- Setup ---
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

PROJECT_DIR = Path(__file__).parent
PDF_DIR = PROJECT_DIR / "pdf_downloads"
CHECKPOINT_PATH = PROJECT_DIR / "checkpoint.json"
LOG_PATH = PROJECT_DIR / "pipeline.log"
EXCEL_PATH = Path(r"E:\HVAC_PDF_search\Unitary_pdf_manual_search_0519_lite.xlsx")
MANIFEST_PATH = PROJECT_DIR / "manifest.json"

PDF_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# --- HTTP Session ---
session = requests.Session()
session.headers.update(
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }
)
session.timeout = 30  # Global timeout for all requests


# ═══════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════


def model_token(model: str) -> str:
    """RACG2T180AC -> RACG2T"""
    if not model:
        return "UNKNOWN"
    m = re.match(r"^([A-Z]+\d+[A-Z]?)", str(model))
    return m.group(1) if m else model[:10]


def model_series(model: str) -> str:
    """RACG2T180AC -> RACG, AD15 -> AD"""
    if not model:
        return "UNKNOWN"
    m = re.match(r"^([A-Z]+)", str(model))
    return m.group(1) if m else "UNKNOWN"


def normalize_brand(brand_label: str) -> str:
    if not brand_label:
        return "UNKNOWN"
    b = str(brand_label).upper()
    for key in [
        "JOHNSON CONTROLS",
        "CARRIER",
        "BRYANT",
        "LENNOX",
        "BOSCH",
        "RUUD",
        "RHEEM",
        "YORK",
    ]:
        if key in b:
            return key
    return b


def safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", name)[:200]


def fetch_url(url: str, timeout: int = 30) -> tuple[int, str]:
    """Fetch URL, return (status_code, text)."""
    try:
        r = session.get(url, timeout=timeout, allow_redirects=True)
        r.encoding = "utf-8"
        return r.status_code, r.text
    except Exception as e:
        return -1, str(e)


def download_pdf(url: str, output_path: Path, timeout: int = 120) -> bool:
    """Download PDF, verify it's valid. Returns True on success."""
    if output_path.exists() and output_path.stat().st_size > 1000:
        return True  # already downloaded
    try:
        r = session.get(url, timeout=timeout)
        if r.status_code == 200 and len(r.content) > 1000:
            if r.content[:4] == b"%PDF":
                output_path.write_bytes(r.content)
                log.info(
                    f"  Downloaded: {output_path.name} ({len(r.content) / 1024:.0f} KB)"
                )
                return True
        return False
    except Exception as e:
        log.warning(f"  Download error: {url[:100]} — {e}")
        return False


def web_search_pdfs(query: str, brand: str = "", max_results: int = 5) -> list[dict]:
    """Search web for PDFs. Returns [{url, title, snippet}]."""
    results = []
    try:
        with DDGS(timeout=15) as ddgs:
            for r in ddgs.text(f"{query}", max_results=max_results):
                url = r.get("href", "")
                title = r.get("title", "")
                body = r.get("body", "")
                if url.lower().endswith(".pdf"):
                    results.append({"url": url, "title": title, "snippet": body[:200]})
                elif any(
                    d in url.lower()
                    for d in [
                        "files.myrheem.com",
                        "carriercca.com/pdf",
                        "docs.johnsoncontrols.com",
                        "cgproducts.johnsoncontrols.com",
                        "lennox.com/dA/",
                        "bosch-homecomfort.com/us/media",
                        "hvacnavigator.com",
                        "bryant.com",
                        "york.com",
                    ]
                ):
                    results.append({"url": url, "title": title, "snippet": body[:200]})
    except Exception as e:
        log.warning(f"  Web search error: {e}")
    return results


def extract_pdf_links_from_html(html: str) -> list[dict]:
    """Extract PDF links from HTML (DocumentURL JSON + href patterns)."""
    results = []
    # Pattern 1: DocumentURL JSON (Rheem style)
    doc_urls = re.findall(r'"DocumentURL"\s*:\s*"(https?://[^"]+\.pdf)"', html)
    doc_types = re.findall(r'"documentType"\s*:\s*"([^"]+)"', html)
    for i, url in enumerate(doc_urls):
        dtype = doc_types[i] if i < len(doc_types) else "PDF"
        results.append({"url": url, "source": f"DocumentURL_{dtype}"})

    # Pattern 2: Generic href
    href_pdfs = re.findall(r'href\s*=\s*"(https?://[^"]+\.pdf)"', html)
    for url in href_pdfs:
        if url not in [r["url"] for r in results]:
            results.append({"url": url, "source": "href"})

    # Pattern 3: data-url/src
    for pattern in [
        r'data-url\s*=\s*"(https?://[^"]+\.pdf)"',
        r'data-src\s*=\s*"(https?://[^"]+\.pdf)"',
    ]:
        for url in re.findall(pattern, html):
            if url not in [r["url"] for r in results]:
                results.append({"url": url, "source": "data_attr"})
    return results


def match_sitemap_urls(
    xml_text: str, tokens: list[str], product_pattern: str = "/product/"
) -> list[str]:
    """Find product page URLs in sitemap XML matching model tokens."""
    all_urls = re.findall(r"<loc>([^<]+)</loc>", xml_text)
    matches = []
    for url in all_urls:
        url_upper = url.upper()
        for token in tokens:
            if token.upper() in url_upper and product_pattern in url.lower():
                matches.append(url)
                break
    return list(dict.fromkeys(matches))  # deduplicate preserving order


# ═══════════════════════════════════════════════════════
# BRAND-SPECIFIC SEARCH STRATEGIES
# ═══════════════════════════════════════════════════════


def search_rheem(entry: dict) -> list[dict]:
    """Rheem: sitemap → product page → DocumentURL. Already ran, retry missing."""
    results = []
    models = entry.get("models", [])
    series = entry.get("series", "")
    brand = entry.get("brand", "RHEEM")

    # Try Rheem sitemap
    sitemaps = [
        "https://www.rheem.com/wp-sitemap-products-1.xml",
        "https://www.rheem.com/wp-sitemap-products-2.xml",
    ]
    for sm_url in sitemaps:
        code, xml_text = fetch_url(sm_url)
        if code != 200:
            continue
        tokens = [model_token(m) for m in models] + [series]
        product_urls = match_sitemap_urls(xml_text, list(set(tokens)))
        if product_urls:
            log.info(f"  [{brand}/{series}] Found {len(product_urls)} product page(s)")
            for purl in product_urls[:1]:  # one product page is enough
                _, html = fetch_url(purl)
                if html:
                    pdf_links = extract_pdf_links_from_html(html)
                    for pl in pdf_links:
                        results.append(
                            {
                                "url": pl["url"],
                                "source": f"sitemap_{pl['source']}",
                                "brand": brand,
                            }
                        )
            if results:
                return results

    # Fallback: web search
    log.info(f"  [{brand}/{series}] Web search fallback")
    for model in models[:2]:
        token = model_token(model)
        for q in [
            f"{brand} {token} specification sheet pdf",
            f"{brand} {token} product data pdf",
            f"site:files.myrheem.com {token}",
        ]:
            for r in web_search_pdfs(q, brand):
                if r["url"] not in [x["url"] for x in results]:
                    results.append(
                        {"url": r["url"], "source": "web_search", "brand": brand}
                    )
    return results


def search_ruud(entry: dict) -> list[dict]:
    """Ruud: commercial product pages + Rheem cross-reference + web search."""
    results = []
    models = entry.get("models", [])
    series = entry.get("series", "")
    brand = "RUUD"

    # Strategy 1: Try Rheem sitemap for same series (sister brand, shared docs)
    for sm_url in ["https://www.rheem.com/wp-sitemap-products-1.xml"]:
        code, xml = fetch_url(sm_url)
        if code == 200:
            tokens = [model_token(m) for m in models] + [series]
            product_urls = match_sitemap_urls(xml, list(set(tokens)))
            for purl in product_urls[:1]:
                _, html = fetch_url(purl)
                if html:
                    for pl in extract_pdf_links_from_html(html):
                        # Rheem PDF but might mention Ruud
                        results.append(
                            {
                                "url": pl["url"],
                                "source": "rheem_sitemap_cross",
                                "brand": brand,
                            }
                        )

    # Strategy 2: Try known Ruud product page URL patterns
    ruud_patterns = {
        "RACCYB": "/product/RACCYB-Renaissance-Line-Achiever-Series-Packaged-AC",
        "RACDZT": "/product/ruud-commercial-package-air-conditioner-RACDZT",
        "RACCYC": "/product/RACCYC-Renaissance-Line-Achiever-Plus-Series-Packaged-AC",
        "RACAYB": "/product/RACAYB-ruud-endeavorline-packagedairconditioner",
        "RACG2T": "/product/RuudRACG2T Renaissance Line Packaged Air Conditioner",
        "RLRL": "/product/ruud-commercial-package-ac-rlrl-c-rlrl-h-15-20-ton",
        "RLNL": "/product/ruud-commercial-package-ac-rlnl-b-rlnl-g-rlnl-h-15-20-ton",
        "RKNL": "/product/ruud-commercial-package-ac-rlkn-b-6-ton",
        "RACCZR": "/product/ruud-achiever-commercial-package-air-conditioner-RACCZRZT",
    }
    for key, path in ruud_patterns.items():
        if key == series or key in str(models):
            encoded = urllib_quote_path(path)
            for url_base in ["https://www.ruud.com", "https://www.rheem.com"]:
                _, html = fetch_url(f"{url_base}{encoded}")
                if html:
                    for pl in extract_pdf_links_from_html(html):
                        results.append(
                            {
                                "url": pl["url"],
                                "source": "ruud_product_page",
                                "brand": brand,
                            }
                        )

    # Strategy 3: Web search
    if not results:
        for model in models[:2]:
            token = model_token(model)
            for q in [
                f"Ruud {token} specification sheet pdf",
                f"Ruud {token} technical guide pdf",
                f"site:files.myrheem.com {token}",
            ]:
                for r in web_search_pdfs(q, "Ruud"):
                    if r["url"] not in [x["url"] for x in results]:
                        results.append(
                            {"url": r["url"], "source": "web_search", "brand": brand}
                        )

    return results


def urllib_quote_path(path: str) -> str:
    """URL-encode a path while preserving slashes."""
    import urllib.parse

    return urllib.parse.quote(path, safe="/")


def search_bosch(entry: dict) -> list[dict]:
    """Bosch: public CDN directory browsing + web search."""
    results = []
    models = entry.get("models", [])
    series = entry.get("series", "")
    brand = "BOSCH"

    token = model_token(models[0]) if models else series

    # Strategy 1: Try known CDN paths

    # Strategy 2: Web search for bosch-homecomfort.com PDFs
    for model in models[:2]:
        for q in [
            f"site:bosch-homecomfort.com {token} pdf",
            f"Bosch {token} engineering submittal pdf",
            f"Bosch {token} product specification pdf",
        ]:
            for r in web_search_pdfs(q, "Bosch"):
                results.append(
                    {"url": r["url"], "source": "web_search", "brand": brand}
                )

    return results


def search_carrier(entry: dict) -> list[dict]:
    """Carrier: carriercca.com CDN + sitemap + web search."""
    results = []
    models = entry.get("models", [])
    series = entry.get("series", "")
    brand = "CARRIER"

    token = model_token(models[0]) if models else series
    model_full = models[0] if models else ""

    # Strategy 1: Try carriercca.com CDN patterns
    cdn_patterns = [
        f"https://www.carriercca.com/pdf/products_pdf/{model_full}_Product_Data.pdf",
        f"https://www.carriercca.com/pdf/products_pdf/{token}_Product_Data.pdf",
        f"https://www.carriercca.com/pdf/products_pdf/{series}_Product_Data.pdf",
        f"https://www.carriercca.com/pdf/products_pdf/{token}_PD.pdf",
        f"https://www.carriercca.com/pdf/products_pdf/{token}_Technical_Guide.pdf",
    ]
    for cdn_url in cdn_patterns:
        try:
            r = session.head(cdn_url, timeout=15)
            if r.status_code == 200:
                results.append(
                    {"url": cdn_url, "source": "carriercca_cdn", "brand": brand}
                )
                break
        except Exception:
            pass

    # Strategy 2: Try Carrier sitemap
    for sm_url in ["https://www.carrier.com/residential/en/us/sitemap.xml"]:
        code, xml = fetch_url(sm_url)
        if code == 200:
            purls = match_sitemap_urls(xml, [token, series], "/products/")
            for purl in purls[:1]:
                _, html = fetch_url(purl)
                if html:
                    results.extend(
                        [
                            {
                                "url": pl["url"],
                                "source": "carrier_sitemap",
                                "brand": brand,
                            }
                            for pl in extract_pdf_links_from_html(html)
                        ]
                    )

    # Strategy 3: Web search
    if not results:
        for model in models[:2]:
            for q in [
                f"Carrier {token} product data pdf",
                f"Carrier {model_full} technical guide pdf",
                f"site:carriercca.com {token} pdf",
            ]:
                for r in web_search_pdfs(q, "Carrier"):
                    results.append(
                        {"url": r["url"], "source": "web_search", "brand": brand}
                    )

    return results


def search_bryant(entry: dict) -> list[dict]:
    """Bryant: Carrier sister brand, similar approach."""
    results = []
    models = entry.get("models", [])
    token = model_token(models[0]) if models else entry.get("series", "")

    # Web search primary (Bryant doesn't have a known CDN)
    for model in models[:2]:
        for q in [
            f"Bryant {token} product data pdf",
            f"Bryant {model} specification sheet pdf",
            f"site:bryant.com {token} pdf",
        ]:
            for r in web_search_pdfs(q, "Bryant"):
                results.append(
                    {"url": r["url"], "source": "web_search", "brand": "BRYANT"}
                )
    return results


def search_lennox(entry: dict) -> list[dict]:
    """Lennox: sitemap + web search."""
    results = []
    models = entry.get("models", [])
    token = model_token(models[0]) if models else entry.get("series", "")

    # Try Lennox sitemap
    code, xml = fetch_url("https://www.lennox.com/sitemap.xml")
    if code == 200:
        purls = match_sitemap_urls(xml, [token, entry.get("series", "")], "/products/")
        for purl in purls[:1]:
            _, html = fetch_url(purl)
            if html:
                results.extend(
                    [
                        {
                            "url": pl["url"],
                            "source": "lennox_sitemap",
                            "brand": "LENNOX",
                        }
                        for pl in extract_pdf_links_from_html(html)
                    ]
                )

    # Web search fallback
    if not results:
        for model in models[:2]:
            for q in [
                f"Lennox {token} product specification pdf",
                f"Lennox {model} technical guide pdf",
                f"site:lennox.com {token} pdf",
            ]:
                for r in web_search_pdfs(q, "Lennox"):
                    results.append(
                        {"url": r["url"], "source": "web_search", "brand": "LENNOX"}
                    )

    return results


def search_jci(entry: dict) -> list[dict]:
    """Johnson Controls: docs.johnsoncontrols.com + web search."""
    results = []
    models = entry.get("models", [])
    series = entry.get("series", "")
    token = model_token(models[0]) if models else series

    # Strategy 1: Try JCI documentation portal
    for doc_url in [
        f"https://docs.johnsoncontrols.com/ductedsystems/api/khub/documents?search={token}",
        f"https://docs.johnsoncontrols.com/ductedsystems/api/khub/documents?search={series}",
    ]:
        try:
            r = session.get(doc_url, timeout=20)
            if r.status_code == 200 and r.text:
                # Try to parse JSON response
                try:
                    data = r.json()
                    if isinstance(data, list):
                        for doc in data:
                            doc_id = doc.get("id", "")
                            if doc_id:
                                pdf_url = f"https://docs.johnsoncontrols.com/ductedsystems/api/khub/documents/{doc_id}/content"
                                results.append(
                                    {
                                        "url": pdf_url,
                                        "source": "jci_api",
                                        "brand": "JCI",
                                    }
                                )
                except json.JSONDecodeError:
                    pass
        except Exception:
            pass

    # Strategy 2: Web search
    for model in models[:2]:
        for q in [
            f"Johnson Controls {token} technical guide pdf",
            f"site:docs.johnsoncontrols.com {token}",
            f"site:cgproducts.johnsoncontrols.com {token} pdf",
            f"Johnson Controls {series} product data pdf",
        ]:
            for r in web_search_pdfs(q, "Johnson Controls"):
                if r["url"] not in [x["url"] for x in results]:
                    results.append(
                        {"url": r["url"], "source": "web_search", "brand": "JCI"}
                    )

    return results


def search_york(entry: dict) -> list[dict]:
    """York: same as JCI approach + york.com sitemap."""
    results = []
    models = entry.get("models", [])
    token = model_token(models[0]) if models else entry.get("series", "")

    # Try York sitemap
    code, xml = fetch_url("https://www.york.com/sitemap.xml")
    if code == 200:
        purls = match_sitemap_urls(xml, [token], "/")
        for purl in purls[:1]:
            _, html = fetch_url(purl)
            if html:
                results.extend(
                    [
                        {"url": pl["url"], "source": "york_sitemap", "brand": "YORK"}
                        for pl in extract_pdf_links_from_html(html)
                    ]
                )

    # Also try JCI sources (York is a JCI brand)
    results.extend([{**r, "brand": "YORK"} for r in search_jci(entry)])

    # Web search
    if not results:
        for model in models[:2]:
            for q in [
                f"York {token} technical guide pdf",
                f"site:york.com {token} pdf",
                f"York {model} specification pdf",
            ]:
                for r in web_search_pdfs(q, "York"):
                    results.append(
                        {"url": r["url"], "source": "web_search", "brand": "YORK"}
                    )

    return results


# Brand search function dispatch
SEARCH_FUNCTIONS = {
    "RHEEM": search_rheem,
    "RUUD": search_ruud,
    "BOSCH": search_bosch,
    "CARRIER": search_carrier,
    "BRYANT": search_bryant,
    "LENNOX": search_lennox,
    "JOHNSON CONTROLS": search_jci,
    "YORK": search_york,
}


# ═══════════════════════════════════════════════════════
# PDF VERIFICATION (pdfplumber)
# ═══════════════════════════════════════════════════════


def verify_pdf(pdf_path: Path, entry: dict) -> dict:
    """Verify PDF content against manifest entry. Returns verdict dict."""
    try:
        import pdfplumber
    except ImportError:
        return {
            "verdict": "Not verified",
            "score": 0,
            "evidence": "pdfplumber not installed",
        }

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            texts = []
            for page in pdf.pages[:50]:
                t = page.extract_text()
                if t:
                    texts.append(t)
        full_text = "\n".join(texts)
    except Exception as e:
        return {
            "verdict": "Not found",
            "score": 0,
            "evidence": f"PDF extraction failed: {e}",
        }

    if len(full_text) < 100:
        return {"verdict": "Not found", "score": 0, "evidence": "PDF text too short"}

    text_upper = full_text.upper()
    brand = entry.get("brand", "")
    series = entry.get("series", "")
    models = entry.get("models", [])
    refs = entry.get("refrigerants", [])
    capacities = entry.get("capacities", [])

    evidence = []
    score = 0

    # Brand match (5 pts)
    brand_aliases = {
        "RHEEM": ["RHEEM", "RUUD"],
        "RUUD": ["RUUD", "RHEEM"],
        "CARRIER": ["CARRIER", "BRYANT"],
        "BRYANT": ["BRYANT", "CARRIER"],
        "LENNOX": ["LENNOX"],
        "BOSCH": ["BOSCH"],
        "YORK": ["YORK", "JOHNSON CONTROLS", "JCI"],
        "JOHNSON CONTROLS": ["JOHNSON CONTROLS", "JCI", "YORK", "CHOICE"],
    }
    for alias in brand_aliases.get(brand, [brand]):
        if alias in text_upper:
            evidence.append(f"Brand: {alias}")
            score += 5
            break
    if score == 0:
        evidence.append(f"Brand NOT found: {brand}")

    # Model match (5 pts)
    model_matched = False
    for model in models:
        if model and model.upper() in text_upper:
            evidence.append(f"Model: {model}")
            score += 5
            model_matched = True
            break
    if not model_matched:
        if series and series.upper() in text_upper:
            evidence.append(f"Series: {series}")
            score += 3
            model_matched = True
    if not model_matched:
        evidence.append(f"Model/Series NOT found: {models}")

    # Refrigerant match (5 pts)
    ref_matched = False
    for ref in refs:
        if not ref:
            continue
        for variant in [ref, ref.replace("-", ""), ref.replace("-", " ")]:
            if variant.upper() in text_upper:
                evidence.append(f"Refrigerant: {ref}")
                score += 5
                ref_matched = True
                break
        if ref_matched:
            break
    if not ref_matched:
        evidence.append(f"Refrigerant NOT confirmed: {refs}")

    # Capacity match (5 pts if found)
    cap_matched = False
    for cap in capacities:
        if not cap:
            continue
        nums = re.findall(r"(\d+)", str(cap))
        for n in nums[:2]:
            if n in full_text:
                evidence.append(f"Capacity: {n} BTU")
                score += 5
                cap_matched = True
                break
        if cap_matched:
            break

    # EER/IEER (5 pts bonus if any performance data found)
    if re.search(r"EER\s*[:\s]*\d+\.?\d*", text_upper) or re.search(
        r"IEER\s*[:\s]*\d+\.?\d*", text_upper
    ):
        evidence.append("Performance data present")
        score += 5

    # Verdict
    if score >= 15:
        verdict = "High confidence"
    elif score >= 8:
        verdict = "Needs human confirmation"
    elif score >= 3:
        verdict = "Low confidence"
    else:
        verdict = "Not found"

    return {"verdict": verdict, "score": score, "evidence": "; ".join(evidence)}


# ═══════════════════════════════════════════════════════
# CORE PIPELINE
# ═══════════════════════════════════════════════════════


def load_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["sheet1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows, headers, wb, ws


def build_manifest(rows):
    manifest = {}
    for i, r in enumerate(rows):
        brand = normalize_brand(str(r.get("Brand_Scope_Label", "")))
        model = str(r.get("ModelNumberUle", ""))
        series = model_series(model)
        mid = f"{brand}__{series}"

        if mid not in manifest:
            manifest[mid] = {
                "id": mid,
                "brand": brand,
                "series": series,
                "pdf_filename": None,
                "pdf_url": None,
                "status": "searching",
                "verdict": None,
                "score": 0,
                "evidence": "",
                "row_indices": [],
                "models": [],
                "reference_ids": [],
                "queries": [],
                "capacities": [],
                "refrigerants": [],
            }
        m = manifest[mid]
        m["row_indices"].append(i)
        if model not in m["models"]:
            m["models"].append(model)
        rid = str(r.get("ReferenceId", ""))
        if rid and rid not in m["reference_ids"]:
            m["reference_ids"].append(rid)
        q = str(r.get("Suggested_Manual_Query", ""))
        if q and q not in m["queries"]:
            m["queries"].append(q)
        cap = str(r.get("CoolingCapacity95FSearchULE", ""))
        if cap and cap not in m["capacities"]:
            m["capacities"].append(cap)
        ref = str(r.get("RefrigerantTypeUle", ""))
        if ref and ref not in m["refrigerants"]:
            m["refrigerants"].append(ref)
    return manifest


def load_checkpoint():
    if CHECKPOINT_PATH.exists():
        with open(CHECKPOINT_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"processed": [], "failed": [], "started_at": None}


def save_checkpoint(manifest):
    processed = [mid for mid, m in manifest.items() if m.get("status") == "found"]
    failed = [mid for mid, m in manifest.items() if m.get("status") == "not_found"]
    cp = {
        "processed": processed,
        "failed": failed,
        "started_at": datetime.now(UTC).isoformat(),
        "total_manuals": len(manifest),
        "found_count": len(processed),
    }
    with open(CHECKPOINT_PATH, "w", encoding="utf-8") as f:
        json.dump(cp, f, indent=2, ensure_ascii=False)


def write_excel_results(manifest):
    rows, headers, wb, ws = load_excel()

    # Ensure output columns
    for col_name in ["pdf_filename", "Match_Status", "Score", "Evidence"]:
        if col_name not in headers:
            col_idx = len(headers) + 1
            ws.cell(row=1, column=col_idx, value=col_name)
            headers.append(col_name)

    pdf_col = headers.index("pdf_filename") + 1
    match_col = headers.index("Match_Status") + 1
    score_col = headers.index("Score") + 1
    ev_col = headers.index("Evidence") + 1
    url_col = headers.index("pdf_link") + 1 if "pdf_link" in headers else -1

    for mid, m in manifest.items():
        for ri in m.get("row_indices", []):
            erow = ri + 2
            if m.get("pdf_filename"):
                ws.cell(row=erow, column=pdf_col, value=m["pdf_filename"])
            if m.get("verdict"):
                ws.cell(row=erow, column=match_col, value=m["verdict"])
            if m.get("score"):
                ws.cell(row=erow, column=score_col, value=m["score"])
            if m.get("evidence"):
                ws.cell(row=erow, column=ev_col, value=m["evidence"])
            if url_col > 0 and m.get("pdf_url"):
                ws.cell(row=erow, column=url_col, value=m["pdf_url"])

    out_path = EXCEL_PATH.parent / "Unitary_pdf_manual_search_0519_lite_RESULTS.xlsx"
    wb.save(out_path)
    log.info(f"Results Excel saved: {out_path}")
    return out_path


def process_entry(entry: dict) -> bool:
    """Process one manifest entry: search → download → verify. Returns True if PDF found."""
    brand = entry["brand"]
    series = entry["series"]
    mid = entry["id"]
    search_fn = SEARCH_FUNCTIONS.get(brand)

    if not search_fn:
        log.warning(f"  [{mid}] No search function for brand '{brand}'")
        return False

    # Search
    log.info(f"  [{mid}] Searching ({len(entry['models'])} models)")
    candidates = search_fn(entry)

    if not candidates:
        log.info(f"  [{mid}] No candidates found")
        entry["status"] = "not_found"
        entry["verdict"] = "Not found"
        return False

    # Try download from candidates
    for cand in candidates[:5]:  # try up to 5 candidates
        pdf_url = cand["url"]
        ref_id = entry["reference_ids"][0] if entry["reference_ids"] else "0"
        ref_type = (
            (entry.get("refrigerants", ["UNKNOWN"])[0] or "UNKNOWN")
            .replace("/", "-")
            .replace(" ", "-")
        )
        stem = pdf_url.split("/")[-1].replace(".pdf", "")[:80]
        fname = safe_filename(f"{stem}__Ref{ref_id}_{brand}_{series}_{ref_type}.pdf")
        outpath = PDF_DIR / fname

        if download_pdf(pdf_url, outpath):
            # Verify
            result = verify_pdf(outpath, entry)
            entry["pdf_filename"] = fname
            entry["pdf_url"] = pdf_url
            entry["status"] = "found"
            entry["verdict"] = result["verdict"]
            entry["score"] = result["score"]
            entry["evidence"] = result["evidence"]
            log.info(f"  [{mid}] FOUND: {result['verdict']} (score={result['score']})")
            return True

    # All candidates failed
    log.info(f"  [{mid}] Downloads failed for {len(candidates)} candidates")
    entry["status"] = "not_found"
    entry["verdict"] = "Not found"
    return False


def run_pipeline(resume: bool = False):
    """Main pipeline — process all manuals."""
    log.info("=" * 60)
    log.info("HVAC PDF MANUAL SEARCH — AUTONOMOUS PIPELINE")
    log.info(f"Start: {datetime.now(UTC).isoformat()}")
    log.info("=" * 60)

    checkpoint = (
        load_checkpoint()
        if resume
        else {"processed": [], "failed": [], "started_at": None}
    )
    if checkpoint["processed"]:
        log.info(
            f"Resuming: {len(checkpoint['processed'])} already processed, {len(checkpoint['failed'])} failed"
        )

    # Load & build manifest
    rows, _, _, _ = load_excel()
    log.info(f"Loaded {len(rows)} rows from Excel")
    manifest = build_manifest(rows)
    log.info(f"Built manifest: {len(manifest)} unique manuals")

    # Restore previous progress if resuming
    for mid in checkpoint.get("processed", []):
        if mid in manifest:
            manifest[mid]["status"] = "found"

    # Sort by brand priority (easiest → hardest)
    brand_order = [
        "RHEEM",
        "RUUD",
        "BOSCH",
        "CARRIER",
        "BRYANT",
        "LENNOX",
        "YORK",
        "JOHNSON CONTROLS",
    ]
    entries = sorted(
        manifest.values(),
        key=lambda e: (
            brand_order.index(e["brand"]) if e["brand"] in brand_order else 99,
            e["series"],
        ),
    )

    pending = [e for e in entries if e["status"] == "searching"]
    log.info(f"Pending: {len(pending)}/{len(manifest)}")

    # Process each
    found_count = 0
    for i, entry in enumerate(pending):
        brand = entry["brand"]
        series = entry["series"]
        log.info(f"[{i + 1}/{len(pending)}] {brand}/{series}")

        try:
            success = process_entry(entry)
            if success:
                found_count += 1
        except Exception as e:
            log.error(f"  ERROR: {e}")
            entry["status"] = "error"
            entry["evidence"] = str(e)[:200]

        # Checkpoint every 5 entries
        if (i + 1) % 5 == 0:
            save_checkpoint(manifest)
            write_excel_results(manifest)
            log.info(f"  [Checkpoint] {found_count} found so far")

        # Rate limiting
        time.sleep(2)

    # Final save
    save_checkpoint(manifest)
    write_excel_results(manifest)

    # Summary
    total = len(manifest)
    found = sum(1 for m in manifest.values() if m.get("status") == "found")
    high = sum(1 for m in manifest.values() if m.get("verdict") == "High confidence")
    needs_review = sum(
        1 for m in manifest.values() if m.get("verdict") == "Needs human confirmation"
    )
    low = sum(1 for m in manifest.values() if m.get("verdict") == "Low confidence")
    not_found = sum(1 for m in manifest.values() if m.get("verdict") == "Not found")

    log.info("=" * 60)
    log.info("PIPELINE COMPLETE")
    log.info(f"  Total manuals: {total}")
    log.info(f"  PDFs found:    {found}")
    log.info(f"  High confidence:   {high}")
    log.info(f"  Needs review:      {needs_review}")
    log.info(f"  Low confidence:    {low}")
    log.info(f"  Not found:         {not_found}")

    # Brand breakdown
    log.info("  By brand:")
    bs = defaultdict(lambda: {"total": 0, "found": 0})
    for m in manifest.values():
        bs[m["brand"]]["total"] += 1
        if m.get("status") == "found":
            bs[m["brand"]]["found"] += 1
    for b, s in sorted(bs.items(), key=lambda x: -x[1]["total"]):
        log.info(f"    {b}: {s['found']}/{s['total']}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true")
    args = parser.parse_args()
    run_pipeline(resume=args.resume)
