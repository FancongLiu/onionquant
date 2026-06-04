"""
HVAC Product Manual Search & Download Tool
============================================
Reads the Excel task list, groups rows by shared manual,
searches manufacturer portals, and downloads PDFs.
"""

import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from pathlib import Path
from collections import defaultdict
import openpyxl
import re
import json

EXCEL_PATH = r"E:\HVAC_PDF_search\Unitary_pdf_manual_search_0519_lite.xlsx"
PDF_DIR = Path(__file__).parent / "pdf_downloads"
MANUAL_MAP_FILE = Path(__file__).parent / "manual_map.json"


def load_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["sheet1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows, headers


def normalize_query(query):
    """Extract the core model series from a search query.
    e.g. 'JOHNSON CONTROLS AD15 packaged rooftop product data pdf'
      -> ('JOHNSON CONTROLS', 'AD')
    """
    if not query:
        return (None, None)
    parts = (
        query.split(" packaged rooftop")[0].strip()
        if "packaged rooftop" in query
        else query
    )
    # Extract brand and model
    for brand in [
        "JOHNSON CONTROLS",
        "BOSCH",
        "CARRIER CORPORATION",
        "BRYANT HEATING & COOLING SYSTEMS",
        "LENNOX",
        "RUUD",
        "RHEEM",
    ]:
        if brand in parts:
            model_part = parts.replace(brand, "").strip()
            # Extract alphabetic prefix of model (e.g. AD from AD15)
            alpha = re.match(r"^([A-Z]+)", model_part)
            series = alpha.group(1) if alpha else model_part
            return (brand, series)
    return (None, None)


def group_by_manual(rows):
    """
    Group rows that share the same manual.
    Key insight: models with the same alphabetic prefix within the same brand
    often share a single technical guide (e.g. AD15, AD18, AD20, AD25, AD28).
    Also use the pdf_link column to identify already-known manuals.
    """
    # First, group by known pdf_link entries
    pdf_link_groups = defaultdict(list)
    for r in rows:
        if r["pdf_link"] and r["pdf_link"] not in ("untitled", "/"):
            # Normalize the pdf_link (remove viewer/chinese text)
            title = str(r["pdf_link"]).split(" • ")[0].strip()
            pdf_link_groups[title].append(r)
        elif r["pdf_link"] == "untitled" or r["pdf_link"] == "/":
            pdf_link_groups[
                f"_UNKNOWN_{r['Brand_Scope_Label']}_{r['Suggested_Manual_Query'][:40]}"
            ].append(r)

    # For remaining rows (empty pdf_link), group by brand + model series
    brand_series_groups = defaultdict(list)
    for r in rows:
        if r["pdf_link"] and str(r["pdf_link"]).strip() not in ("untitled", "/"):
            continue  # already in a pdf_link group
        brand, series = normalize_query(r["Suggested_Manual_Query"])
        if brand and series:
            brand_series_groups[(brand, series)].append(r)
        else:
            brand_series_groups[
                ("_UNGROUPED", str(r["Suggested_Manual_Query"])[:60])
            ].append(r)

    return pdf_link_groups, brand_series_groups


def main():
    rows, headers = load_excel()
    print(f"Loaded {len(rows)} rows from Excel")
    print(f"PDF storage: {PDF_DIR}")
    print()

    pdf_link_groups, brand_series_groups = group_by_manual(rows)

    # Report: Known manuals from pdf_link column
    print("=" * 80)
    print("KNOWN MANUALS (from pdf_link column)")
    print("=" * 80)
    for title, group_rows in sorted(pdf_link_groups.items(), key=lambda x: -len(x[1])):
        models = [r["ModelNumberUle"][:20] for r in group_rows]
        statuses = set(r["ModelStatusId"] for r in group_rows)
        print(f"  [{len(group_rows)} models] {title[:100]}")
        print(f"    Statuses: {statuses}")
        print(f"    Models: {', '.join(models[:5])}{'...' if len(models) > 5 else ''}")
        print()

    # Report: Inferred groupings
    print("=" * 80)
    print("INFERRED GROUPINGS (by brand + model series, no pdf_link yet)")
    print("=" * 80)
    total_inferred = sum(len(v) for v in brand_series_groups.values())
    print(f"  Total rows still needing manuals: {total_inferred}")
    print(f"  Unique groups: {len(brand_series_groups)}")
    print()
    # Show top groups
    for (brand, series), group_rows in sorted(
        brand_series_groups.items(), key=lambda x: -len(x[1])
    )[:30]:
        queries = list(set(str(r["Suggested_Manual_Query"])[:60] for r in group_rows))
        print(f"  [{len(group_rows)} rows] {brand} / {series}")
        if len(queries) <= 3:
            for q in queries:
                print(f"    -> {q}")
        else:
            print(f"    -> {queries[0]} ... ({len(queries)} unique queries)")
    print()

    # Save mapping for future use
    manual_map = {
        "known_manuals": {
            title: {
                "rows": len(group_rows),
                "models": [r["ModelNumberUle"] for r in group_rows],
                "reference_ids": [r["ReferenceId"] for r in group_rows],
                "queries": [r["Suggested_Manual_Query"] for r in group_rows],
                "brand": group_rows[0]["Brand_Scope_Label"],
                "status": list(set(r["ModelStatusId"] for r in group_rows)),
            }
            for title, group_rows in pdf_link_groups.items()
        },
        "inferred_groups": {
            f"{brand}/{series}": {
                "rows": len(group_rows),
                "models": [r["ModelNumberUle"] for r in group_rows],
                "queries": [r["Suggested_Manual_Query"] for r in group_rows],
            }
            for (brand, series), group_rows in brand_series_groups.items()
        },
    }
    with open(MANUAL_MAP_FILE, "w", encoding="utf-8") as f:
        json.dump(manual_map, f, indent=2, ensure_ascii=False)
    print(f"Manual mapping saved to: {MANUAL_MAP_FILE}")


if __name__ == "__main__":
    main()
